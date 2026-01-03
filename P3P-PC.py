import openpyxl
import csv
import os

# Caminho do arquivo Excel
excel_file = "P3P-PC.xlsx"  # Substitua pelo caminho do seu arquivo Excel

# Cria a pasta 'Text' se não existir
output_dir = "Texts"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Carrega o arquivo Excel
workbook = openpyxl.load_workbook(excel_file, read_only=True)

# Itera sobre todas as abas
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Define o caminho do arquivo TSV
    tsv_file = os.path.join(output_dir, f"{sheet_name}.txt")
    
    # Escreve diretamente como TSV
    with open(tsv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t', lineterminator='\n')
        for row in sheet.iter_rows(values_only=True):
            # Converte valores para string e lida com None
            row = ['' if cell is None else str(cell) for cell in row]
            writer.writerow(row)

# Fecha o workbook
workbook.close()